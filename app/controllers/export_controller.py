from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import csv
from io import BytesIO, StringIO
from flask import Response
from datetime import date, timedelta
from app.controllers.insumo_controller import InsumoController, MovimentacaoInsumoController

class RelatoriosInsumosController:
    
    @staticmethod
    def consumo(data_inicio, data_fim, categoria=None):
        """Gera dados do relatório de consumo"""
        # Buscar movimentações de saída no período
        movimentacoes = MovimentacaoInsumoController.listar_movimentacoes(
            tipo='Saída - Uso',
            data_inicio=data_inicio,
            data_fim=data_fim
        )
        
        # Agrupar por insumo
        consumo_por_insumo = {}
        for mov in movimentacoes:
            if categoria and mov.insumo.categoria != categoria:
                continue
                
            nome_insumo = mov.insumo.nome
            if nome_insumo not in consumo_por_insumo:
                consumo_por_insumo[nome_insumo] = {
                    'insumo': mov.insumo,
                    'total_consumido': 0,
                    'movimentacoes': []
                }
            
            consumo_por_insumo[nome_insumo]['total_consumido'] += mov.quantidade
            consumo_por_insumo[nome_insumo]['movimentacoes'].append(mov)
        
        # Ordenar por consumo
        consumo_ordenado = sorted(
            consumo_por_insumo.items(), 
            key=lambda x: x[1]['total_consumido'], 
            reverse=True
        )
        
        return {
            'consumo_por_insumo': consumo_por_insumo,
            'consumo_ordenado': consumo_ordenado,
            'total_itens': len(consumo_por_insumo),
            'total_quantidade': sum(dados['total_consumido'] for dados in consumo_por_insumo.values())
        }
    
    @staticmethod
    def cobertura():
        """Gera relatório de cobertura de estoque"""
        insumos = InsumoController.listar_todos(ativo=True)
        
        # Calcular cobertura baseada no consumo dos últimos 30 dias
        data_fim = date.today()
        data_inicio = data_fim - timedelta(days=30)
        
        relatorio_cobertura = []
        
        for insumo in insumos:
            # Buscar movimentações do insumo nos últimos 30 dias
            movimentacoes = MovimentacaoInsumoController.listar_movimentacoes(
                insumo_id=insumo.id_insumo,
                tipo='Saída - Uso',
                data_inicio=data_inicio,
                data_fim=data_fim
            )
            
            total_consumido = sum(mov.quantidade for mov in movimentacoes)
            media_diaria = total_consumido / 30 if total_consumido > 0 else 0
            dias_cobertura = (insumo.quantidade_atual / media_diaria) if media_diaria > 0 else 999
            
            relatorio_cobertura.append({
                'insumo': insumo,
                'total_consumido_30d': total_consumido,
                'media_diaria': media_diaria,
                'dias_cobertura': dias_cobertura,
                'status': 'crítico' if dias_cobertura < 7 else 'atenção' if dias_cobertura < 30 else 'ok'
            })
        
        # Ordenar por dias de cobertura (menor primeiro)
        relatorio_cobertura.sort(key=lambda x: x['dias_cobertura'])
        
        return relatorio_cobertura
    
    @staticmethod
    def vencimentos(dias=30):
        """Gera relatório de vencimentos próximos"""
        insumos = InsumoController.listar_todos(vencimento_dias=dias, ativo=True)
        
        relatorio_vencimentos = []
        hoje = date.today()
        
        for insumo in insumos:
            if insumo.data_validade:
                dias_restantes = (insumo.data_validade - hoje).days
                status = 'vencido' if dias_restantes < 0 else 'crítico' if dias_restantes <= 7 else 'atenção'
                
                relatorio_vencimentos.append({
                    'insumo': insumo,
                    'dias_restantes': dias_restantes,
                    'status': status
                })
        
        # Ordenar por dias restantes
        relatorio_vencimentos.sort(key=lambda x: x['dias_restantes'])
        
        return relatorio_vencimentos
    
    @staticmethod
    def abaixo_minimo():
        """Gera relatório de insumos abaixo do mínimo"""
        insumos = InsumoController.listar_todos(abaixo_minimo=True, ativo=True)
        
        relatorio_minimo = []
        
        for insumo in insumos:
            diferenca = insumo.quantidade_atual - insumo.quantidade_minima
            percentual = (insumo.quantidade_atual / insumo.quantidade_minima * 100) if insumo.quantidade_minima > 0 else 0
            
            relatorio_minimo.append({
                'insumo': insumo,
                'diferenca': diferenca,
                'percentual': percentual,
                'criticidade': 'crítico' if percentual < 50 else 'atenção'
            })
        
        # Ordenar por percentual (menor primeiro)
        relatorio_minimo.sort(key=lambda x: x['percentual'])
        
        return relatorio_minimo
    
    @staticmethod
    def exportar(tipo_relatorio, formato, **kwargs):
        """Exporta relatórios em diferentes formatos"""
        
        if tipo_relatorio == 'consumo':
            return RelatoriosInsumosController._exportar_consumo(formato, **kwargs)
        elif tipo_relatorio == 'cobertura':
            return RelatoriosInsumosController._exportar_cobertura(formato, **kwargs)
        elif tipo_relatorio == 'vencimentos':
            return RelatoriosInsumosController._exportar_vencimentos(formato, **kwargs)
        elif tipo_relatorio == 'abaixo_minimo':
            return RelatoriosInsumosController._exportar_abaixo_minimo(formato, **kwargs)
        else:
            raise ValueError("Tipo de relatório inválido")
    
    @staticmethod
    def _exportar_consumo(formato, data_inicio, data_fim, categoria=None):
        """Exporta relatório de consumo"""
        dados = RelatoriosInsumosController.consumo(data_inicio, data_fim, categoria)
        
        if formato == 'csv':
            return RelatoriosInsumosController._export_consumo_csv(dados, data_inicio, data_fim)
        elif formato == 'excel':
            return RelatoriosInsumosController._export_consumo_excel(dados, data_inicio, data_fim)
        elif formato == 'pdf':
            return RelatoriosInsumosController._export_consumo_pdf(dados, data_inicio, data_fim)
    
    @staticmethod
    def _exportar_cobertura(formato, **kwargs):
        """Exporta relatório de cobertura"""
        dados = RelatoriosInsumosController.cobertura()
        
        if formato == 'csv':
            return RelatoriosInsumosController._export_cobertura_csv(dados)
        elif formato == 'excel':
            return RelatoriosInsumosController._export_cobertura_excel(dados)
        elif formato == 'pdf':
            return RelatoriosInsumosController._export_cobertura_pdf(dados)
    
    @staticmethod
    def _exportar_vencimentos(formato, dias=30):
        """Exporta relatório de vencimentos"""
        dados = RelatoriosInsumosController.vencimentos(dias)
        
        if formato == 'csv':
            return RelatoriosInsumosController._export_vencimentos_csv(dados)
        elif formato == 'excel':
            return RelatoriosInsumosController._export_vencimentos_excel(dados)
        elif formato == 'pdf':
            return RelatoriosInsumosController._export_vencimentos_pdf(dados)
    
    @staticmethod
    def _exportar_abaixo_minimo(formato, **kwargs):
        """Exporta relatório de abaixo do mínimo"""
        dados = RelatoriosInsumosController.abaixo_minimo()
        
        if formato == 'csv':
            return RelatoriosInsumosController._export_abaixo_minimo_csv(dados)
        elif formato == 'excel':
            return RelatoriosInsumosController._export_abaixo_minimo_excel(dados)
        elif formato == 'pdf':
            return RelatoriosInsumosController._export_abaixo_minimo_pdf(dados)
        
    # ===== EXPORTAÇÕES VENCIMENTOS =====
    @staticmethod
    def _export_vencimentos_csv(dados):
        """Exporta vencimentos em CSV"""
        output = StringIO()
        writer = csv.writer(output)
        
        # Cabeçalho
        writer.writerow(['Insumo', 'Categoria', 'Data_Validade', 'Dias_Restantes', 'Estoque_Atual', 'Unidade', 'Status'])
        
        # Dados
        for item in dados:
            writer.writerow([
                item['insumo'].nome,
                item['insumo'].categoria,
                item['insumo'].data_validade.strftime('%d/%m/%Y') if item['insumo'].data_validade else '',
                item['dias_restantes'],
                float(item['insumo'].quantidade_atual),
                item['insumo'].unidade,
                item['status']
            ])
        
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=vencimentos_insumos_{date.today()}.csv'}
        )
    
    @staticmethod
    def _export_vencimentos_excel(dados):
        """Exporta vencimentos em Excel"""
        buffer = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vencimentos"
        
        # Cabeçalho
        header_fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
        header_font = Font(color='000000', bold=True)
        
        headers = ['Insumo', 'Categoria', 'Data Validade', 'Dias Restantes', 'Estoque Atual', 'Unidade', 'Status']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)  
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        for i, item in enumerate(dados, 1):
            row = i + 1
            
            ws.cell(row=row, column=1, value=item['insumo'].nome)  
            ws.cell(row=row, column=2, value=item['insumo'].categoria)  
            ws.cell(row=row, column=3, value=item['insumo'].data_validade.strftime('%d/%m/%Y') if item['insumo'].data_validade else '') 
            ws.cell(row=row, column=4, value=item['dias_restantes']) 
            ws.cell(row=row, column=5, value=float(item['insumo'].quantidade_atual)) 
            ws.cell(row=row, column=6, value=item['insumo'].unidade)  
            ws.cell(row=row, column=7, value=item['status']) 
        
        # Autofit
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=vencimentos_insumos_{date.today()}.xlsx'}
        )
    
    @staticmethod
    def _export_vencimentos_pdf(dados):
        """Exporta vencimentos em PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        title = Paragraph("Relatório - Vencimentos Próximos", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Dados da tabela
        table_data = [['Insumo', 'Categoria', 'Data Validade', 'Dias Restantes', 'Status']]
        
        for item in dados:
            data_validade = item['insumo'].data_validade.strftime('%d/%m/%Y') if item['insumo'].data_validade else 'Sem data'
            table_data.append([
                item['insumo'].nome,
                item['insumo'].categoria,
                data_validade,
                f"{item['dias_restantes']} dias",
                item['status'].upper()
            ])
        
        # Criar tabela
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.orange),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Rodapé
        story.append(Spacer(1, 30))
        footer = Paragraph(f"Gerado em: {date.today().strftime('%d/%m/%Y')}", styles['Normal'])
        story.append(footer)
        
        doc.build(story)
        buffer.seek(0)
        
        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename=vencimentos_insumos_{date.today()}.pdf'}
        )
    
    # ===== EXPORTAÇÕES CSV =====
    @staticmethod
    def _export_consumo_csv(dados, data_inicio, data_fim):
        """Exporta consumo em CSV"""
        output = StringIO()
        writer = csv.writer(output)
        
        # Cabeçalho
        writer.writerow(['Ranking', 'Insumo', 'Categoria', 'Total_Consumido', 'Unidade', 'Media_Diaria', 'Num_Saidas', 'Estoque_Atual'])
        
        # Dados
        dias_periodo = (data_fim - data_inicio).days + 1
        
        for i, (nome_insumo, dados_insumo) in enumerate(dados['consumo_ordenado'], 1):
            media_diaria = dados_insumo['total_consumido'] / dias_periodo
            writer.writerow([
                i,
                nome_insumo,
                dados_insumo['insumo'].categoria,
                float(dados_insumo['total_consumido']),
                dados_insumo['insumo'].unidade,
                round(media_diaria, 2),
                len(dados_insumo['movimentacoes']),
                float(dados_insumo['insumo'].quantidade_atual)
            ])
        
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=consumo_insumos_{data_inicio}_{data_fim}.csv'}
        )
    
    @staticmethod
    def _export_abaixo_minimo_csv(dados):
        """Exporta abaixo do mínimo em CSV"""
        output = StringIO()
        writer = csv.writer(output)
        
        # Cabeçalho
        writer.writerow(['Insumo', 'Categoria', 'Estoque_Atual', 'Estoque_Minimo', 'Diferenca', 'Percentual', 'Unidade', 'Criticidade'])
        
        # Dados
        for item in dados:
            writer.writerow([
                item['insumo'].nome,
                item['insumo'].categoria,
                float(item['insumo'].quantidade_atual),
                float(item['insumo'].quantidade_minima),
                float(item['diferenca']),
                round(item['percentual'], 1),
                item['insumo'].unidade,
                item['criticidade']
            ])
        
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=insumos_abaixo_minimo_{date.today()}.csv'}
        )
    
    # ===== EXPORTAÇÕES EXCEL =====
    @staticmethod
    def _export_consumo_excel(dados, data_inicio, data_fim):
        """Exporta consumo em Excel"""
        buffer = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consumo de Insumos"
        
        # Cabeçalho
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        headers = ['Ranking', 'Insumo', 'Categoria', 'Total Consumido', 'Unidade', 'Média Diária', 'Nº Saídas', 'Estoque Atual']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header) 
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        dias_periodo = (data_fim - data_inicio).days + 1
        
        for i, (nome_insumo, dados_insumo) in enumerate(dados['consumo_ordenado'], 1):
            media_diaria = dados_insumo['total_consumido'] / dias_periodo
            row = i + 1
            
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=nome_insumo) 
            ws.cell(row=row, column=3, value=dados_insumo['insumo'].categoria)  
            ws.cell(row=row, column=4, value=float(dados_insumo['total_consumido'])) 
            ws.cell(row=row, column=5, value=dados_insumo['insumo'].unidade)  
            ws.cell(row=row, column=6, value=round(media_diaria, 2))  
            ws.cell(row=row, column=7, value=len(dados_insumo['movimentacoes'])) 
            ws.cell(row=row, column=8, value=float(dados_insumo['insumo'].quantidade_atual)) 
        
        # Autofit colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=consumo_insumos_{data_inicio}_{data_fim}.xlsx'}
        )
    
    @staticmethod
    def _export_abaixo_minimo_excel(dados):
        """Exporta abaixo do mínimo em Excel"""
        buffer = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Abaixo do Mínimo"
        
        # Cabeçalho
        header_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        headers = ['Insumo', 'Categoria', 'Estoque Atual', 'Estoque Mínimo', 'Diferença', 'Percentual', 'Unidade', 'Criticidade']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        for i, item in enumerate(dados, 1):
            row = i + 1
            
            ws.cell(row=row, column=1, value=item['insumo'].nome) 
            ws.cell(row=row, column=2, value=item['insumo'].categoria)  
            ws.cell(row=row, column=3, value=float(item['insumo'].quantidade_atual))  
            ws.cell(row=row, column=4, value=float(item['insumo'].quantidade_minima)) 
            ws.cell(row=row, column=5, value=float(item['diferenca']))  
            ws.cell(row=row, column=6, value=round(item['percentual'], 1))  
            ws.cell(row=row, column=7, value=item['insumo'].unidade)  
            ws.cell(row=row, column=8, value=item['criticidade'])  
        
        # Autofit
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=insumos_abaixo_minimo_{date.today()}.xlsx'}
        )
    
    # ===== EXPORTAÇÕES PDF =====
    @staticmethod
    def _export_consumo_pdf(dados, data_inicio, data_fim):
        """Exporta consumo em PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        title = Paragraph("Relatório de Consumo de Insumos", title_style)
        story.append(title)
        
        # Período
        periodo_text = f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
        periodo = Paragraph(periodo_text, styles['Normal'])
        story.append(periodo)
        story.append(Spacer(1, 20))
        
        # Dados da tabela
        table_data = [['Ranking', 'Insumo', 'Categoria', 'Total Consumido', 'Média Diária']]
        
        dias_periodo = (data_fim - data_inicio).days + 1
        
        for i, (nome_insumo, dados_insumo) in enumerate(dados['consumo_ordenado'], 1):
            media_diaria = dados_insumo['total_consumido'] / dias_periodo
            table_data.append([
                f"{i}º",
                nome_insumo,
                dados_insumo['insumo'].categoria,
                f"{dados_insumo['total_consumido']} {dados_insumo['insumo'].unidade}",
                f"{media_diaria:.2f} {dados_insumo['insumo'].unidade}/dia"
            ])
        
        # Criar tabela
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Rodapé
        story.append(Spacer(1, 30))
        footer = Paragraph(f"Gerado em: {date.today().strftime('%d/%m/%Y')}", styles['Normal'])
        story.append(footer)
        
        doc.build(story)
        buffer.seek(0)
        
        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename=consumo_insumos_{data_inicio}_{data_fim}.pdf'}
        )
    
    @staticmethod
    def _export_abaixo_minimo_pdf(dados):
        """Exporta abaixo do mínimo em PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        title = Paragraph("Relatório - Insumos Abaixo do Mínimo", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Dados da tabela
        table_data = [['Insumo', 'Categoria', 'Atual', 'Mínimo', 'Diferença', 'Status']]
        
        for item in dados:
            table_data.append([
                item['insumo'].nome,
                item['insumo'].categoria,
                f"{item['insumo'].quantidade_atual} {item['insumo'].unidade}",
                f"{item['insumo'].quantidade_minima} {item['insumo'].unidade}",
                f"{item['diferenca']:.1f} {item['insumo'].unidade}",
                item['criticidade'].upper()
            ])
        
        # Criar tabela
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.red),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Rodapé
        story.append(Spacer(1, 30))
        footer = Paragraph(f"Gerado em: {date.today().strftime('%d/%m/%Y')}", styles['Normal'])
        story.append(footer)
        
        doc.build(story)
        buffer.seek(0)
        
        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename=insumos_abaixo_minimo_{date.today()}.pdf'}
        )